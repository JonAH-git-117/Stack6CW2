# Standard library imports
import io
import os

# Excel library
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Django imports
from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.http import FileResponse, HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string

# Models
from teams.models import Department, Organisation, Team


def _report_context():
    return {
        'teams': Team.objects.select_related(
            'department',
            'department__organisation',
            'manager'
        ).prefetch_related('members'),
        'departments': Department.objects.annotate(
            team_count=Count('teams')
        ).select_related('organisation', 'dept_head'),
        'teams_without_manager': Team.objects.filter(
            manager__isnull=True
        ).select_related('department'),
        'total_teams': Team.objects.count(),
        'total_departments': Department.objects.count(),
    }


def _escape_pdf_text(value):
    return str(value).replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def _build_simple_pdf(lines):
    text_commands = ["BT", "/F1 12 Tf", "72 770 Td", "16 TL"]
    for line in lines[:42]:
        text_commands.append(f"({_escape_pdf_text(line)}) Tj")
        text_commands.append("T*")
    text_commands.append("ET")
    stream = "\n".join(text_commands).encode("latin-1", errors="replace")

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
    ]

    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{index} 0 obj\n".encode("ascii"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")
    xref_offset = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    return bytes(pdf)


@login_required
def reports_dashboard(request):
    total_teams = Team.objects.count()
    total_departments = Department.objects.count()

    teams_without_manager = Team.objects.filter(
        manager__isnull=True
    ).select_related('department')

    active_teams = Team.objects.filter(status='active').count()
    disabled_teams = Team.objects.filter(status='disabled').count()

    dept_summary = Department.objects.annotate(
        team_count=Count('teams')
    ).select_related('organisation')

    context = {
        'total_teams': total_teams,
        'total_departments': total_departments,
        'teams_without_manager': teams_without_manager,
        'active_teams': active_teams,
        'disabled_teams': disabled_teams,
        'dept_summary': dept_summary,
    }

    return render(request, 'reports/reports_dashboard.html', context)


@login_required
def generate_pdf(request):
    context = _report_context()
    html_string = render_to_string('reports/report.html', context)

    try:
        from weasyprint import HTML
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf = html.write_pdf()
    except Exception:
        lines = [
            "Sky Engineering Teams Report",
            f"Total teams: {context['total_teams']}",
            f"Total departments: {context['total_departments']}",
            f"Teams without managers: {context['teams_without_manager'].count()}",
            "",
            "Departments",
        ]
        for dept in context['departments']:
            lines.append(f"- {dept.department_name}: {dept.team_count} teams")
        lines.append("")
        lines.append("Teams")
        for team in context['teams']:
            manager = team.manager.get_full_name() or team.manager.username if team.manager else "No Manager"
            lines.append(f"- {team.name} | {team.department.department_name} | {manager}")
        pdf = _build_simple_pdf(lines)

    buffer = io.BytesIO(pdf)
    return FileResponse(buffer, as_attachment=True, filename='sky_teams_report.pdf')


@login_required
def generate_excel(request):
    workbook = openpyxl.Workbook()

    sheet1 = workbook.active
    sheet1.title = 'All Teams'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1e3a8a')
    center = Alignment(horizontal='center')

    headers = ['Team Name', 'Department', 'Organisation', 'Manager', 'Status', 'Members']

    for col, header in enumerate(headers, 1):
        cell = sheet1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    teams = Team.objects.select_related(
        'department',
        'department__organisation',
        'manager'
    ).prefetch_related('members')

    for row, team in enumerate(teams, 2):
        manager_name = (
            team.manager.get_full_name() or team.manager.username
            if team.manager else 'No Manager'
        )

        sheet1.cell(row=row, column=1, value=team.name)
        sheet1.cell(row=row, column=2, value=team.department.department_name)
        sheet1.cell(row=row, column=3, value=team.department.organisation.organisation_name)
        sheet1.cell(row=row, column=4, value=manager_name)
        sheet1.cell(row=row, column=5, value=team.get_status_display())
        sheet1.cell(row=row, column=6, value=team.members.count())

    for col in sheet1.columns:
        sheet1.column_dimensions[col[0].column_letter].width = 22

    # Sheet 2
    sheet2 = workbook.create_sheet(title='Department Summary')

    dept_headers = ['Department', 'Organisation', 'Team Count', 'Dept Head']

    for col, header in enumerate(dept_headers, 1):
        cell = sheet2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    departments = Department.objects.annotate(
        team_count=Count('teams')
    ).select_related('organisation', 'dept_head')

    for row, dept in enumerate(departments, 2):
        head_name = (
            dept.dept_head.get_full_name() or dept.dept_head.username
            if dept.dept_head else 'Not Assigned'
        )

        sheet2.cell(row=row, column=1, value=dept.department_name)
        sheet2.cell(row=row, column=2, value=dept.organisation.organisation_name)
        sheet2.cell(row=row, column=3, value=dept.team_count)
        sheet2.cell(row=row, column=4, value=head_name)

    for col in sheet2.columns:
        sheet2.column_dimensions[col[0].column_letter].width = 22

    # Sheet 3
    sheet3 = workbook.create_sheet(title='No Manager')

    no_mgr_headers = ['Team Name', 'Department', 'Status']

    for col, header in enumerate(no_mgr_headers, 1):
        cell = sheet3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill('solid', fgColor='cc0000')
        cell.alignment = center

    teams_no_manager = Team.objects.filter(
        manager__isnull=True
    ).select_related('department')

    for row, team in enumerate(teams_no_manager, 2):
        sheet3.cell(row=row, column=1, value=team.name)
        sheet3.cell(row=row, column=2, value=team.department.department_name)
        sheet3.cell(row=row, column=3, value=team.get_status_display())

    for col in sheet3.columns:
        sheet3.column_dimensions[col[0].column_letter].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = 'attachment; filename="sky_teams_report.xlsx"'

    workbook.save(response)
    return response
